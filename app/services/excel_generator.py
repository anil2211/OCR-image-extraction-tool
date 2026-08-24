from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


class ExcelGenerator:

    @staticmethod
    def generate(
        document: dict[str, Any],
        output_path: str | Path,
    ) -> Path:

        output_path = Path(output_path)

        workbook = Workbook()

        # Remove default worksheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        pages = document.get("pages", [])

        for page in pages:

            page_number = page.get(
                "page_number",
                1
            )

            sheet = workbook.create_sheet(
                title=f"Page {page_number}"
            )

            row = 1

            # ==================================================
            # EXTRACTED TEXT
            # ==================================================

            title_cell = sheet.cell(
                row=row,
                column=1,
                value="Extracted Text"
            )

            title_cell.font = Font(
                bold=True
            )

            row += 1

            text = page.get(
                "text",
                ""
            )

            if text:

                for line in text.splitlines():

                    line = line.strip()

                    if not line:
                        continue

                    cell = sheet.cell(
                        row=row,
                        column=1,
                        value=line
                    )

                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                    row += 1

            row += 2

            # ==================================================
            # TABLES
            # ==================================================

            tables = page.get(
                "tables",
                []
            )

            for table_index, table in enumerate(
                tables,
                start=1
            ):

                table_title = sheet.cell(
                    row=row,
                    column=1,
                    value=f"Table {table_index}"
                )

                table_title.font = Font(
                    bold=True
                )

                row += 1

                rows = (
                    ExcelGenerator.extract_table_rows(
                        table
                    )
                )

                if not rows:
                    row += 1
                    continue

                for table_row in rows:

                    for column_index, value in enumerate(
                        table_row,
                        start=1
                    ):

                        cell = sheet.cell(
                            row=row,
                            column=column_index,
                            value=(
                                ""
                                if value is None
                                else str(value)
                            )
                        )

                        cell.alignment = Alignment(
                            vertical="top",
                            wrap_text=True
                        )

                    row += 1

                row += 2

            # ==================================================
            # COLUMN WIDTH
            # ==================================================

            for column_index in range(
                1,
                sheet.max_column + 1
            ):

                column_letter = (
                    get_column_letter(
                        column_index
                    )
                )

                max_length = 0

                for cell in sheet[
                    column_letter
                ]:

                    if cell.value is None:
                        continue

                    value_length = len(
                        str(cell.value)
                    )

                    max_length = max(
                        max_length,
                        value_length
                    )

                sheet.column_dimensions[
                    column_letter
                ].width = min(
                    max(max_length + 2, 12),
                    60
                )

        # ======================================================
        # FALLBACK
        # ======================================================

        if not workbook.sheetnames:

            sheet = workbook.create_sheet(
                title="Extracted Text"
            )

            sheet["A1"] = (
                "No content was extracted."
            )

        workbook.save(
            output_path
        )

        return output_path

    # ==========================================================
    # TABLE NORMALIZATION
    # ==========================================================

    @staticmethod
    def extract_table_rows(
        table: Any
    ) -> list[list]:

        if table is None:
            return []

        # ------------------------------------------------------
        # Already normalized:
        #
        # [
        #   ["Name", "Age"],
        #   ["Anil", "25"]
        # ]
        # ------------------------------------------------------

        if isinstance(table, list):

            if not table:
                return []

            if all(
                isinstance(row, (list, tuple))
                for row in table
            ):

                return [
                    list(row)
                    for row in table
                ]

        # ------------------------------------------------------
        # Dictionary result
        # ------------------------------------------------------

        if isinstance(table, dict):

            # Common table result keys
            for key in (
                "structure",
                "res",
                "table",
                "table_res",
                "table_result",
            ):

                value = table.get(key)

                if value is None:
                    continue

                nested_rows = (
                    ExcelGenerator.extract_table_rows(
                        value
                    )
                )

                if nested_rows:
                    return nested_rows

            # --------------------------------------------------
            # HTML table
            # --------------------------------------------------

            html = table.get(
                "html"
            )

            if isinstance(html, str):

                rows = (
                    ExcelGenerator.parse_html_table(
                        html
                    )
                )

                if rows:
                    return rows

            # --------------------------------------------------
            # Cells
            # --------------------------------------------------

            cells = table.get(
                "cells"
            )

            if isinstance(cells, list):

                rows = []

                for cell in cells:

                    if isinstance(cell, dict):

                        text = cell.get(
                            "text",
                            ""
                        )

                        rows.append([
                            text
                        ])

                    elif isinstance(cell, str):

                        rows.append([
                            cell
                        ])

                if rows:
                    return rows

        # ------------------------------------------------------
        # String fallback
        # ------------------------------------------------------

        if isinstance(table, str):

            return [
                [line]
                for line in table.splitlines()
                if line.strip()
            ]

        return []

    # ==========================================================
    # SIMPLE HTML TABLE PARSER
    # ==========================================================

    @staticmethod
    def parse_html_table(
        html: str
    ) -> list[list]:

        try:

            from html.parser import HTMLParser

        except ImportError:

            return []

        class TableParser(
            HTMLParser
        ):

            def __init__(self):

                super().__init__()

                self.rows = []

                self.current_row = None

                self.current_cell = None

            def handle_starttag(
                self,
                tag,
                attrs
            ):

                tag = tag.lower()

                if tag == "tr":

                    self.current_row = []

                elif tag in (
                    "td",
                    "th",
                ):

                    self.current_cell = ""

            def handle_data(
                self,
                data
            ):

                if self.current_cell is not None:

                    self.current_cell += data

            def handle_endtag(
                self,
                tag
            ):

                tag = tag.lower()

                if tag in (
                    "td",
                    "th",
                ):

                    if self.current_row is not None:

                        self.current_row.append(
                            self.current_cell.strip()
                        )

                    self.current_cell = None

                elif tag == "tr":

                    if (
                        self.current_row
                        and any(
                            str(value).strip()
                            for value
                            in self.current_row
                        )
                    ):

                        self.rows.append(
                            self.current_row
                        )

                    self.current_row = None

        parser = TableParser()

        parser.feed(html)

        return parser.rows